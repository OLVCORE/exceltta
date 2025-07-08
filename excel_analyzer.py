import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from pathlib import Path
import re

def analyze_excel_file(file_path):
    """
    Analisa uma planilha Excel e extrai informações sobre:
    - Todas as abas
    - Links e hyperlinks
    - Fontes de dados
    - Estrutura geral
    """
    
    print(f"🔍 Analisando planilha: {file_path}")
    print("=" * 60)
    
    # Carregar o workbook
    try:
        wb = load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"❌ Erro ao carregar a planilha: {e}")
        return
    
    analysis = {
        "file_name": Path(file_path).name,
        "file_size_mb": round(Path(file_path).stat().st_size / (1024 * 1024), 2),
        "sheets": [],
        "hyperlinks": [],
        "external_links": [],
        "data_sources": [],
        "formulas": [],
        "charts": [],
        "summary": {}
    }
    
    # Analisar cada aba
    for sheet_name in wb.sheetnames:
        print(f"\n📊 Analisando aba: {sheet_name}")
        
        sheet = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "dimensions": f"{sheet.max_row} linhas x {sheet.max_column} colunas",
            "used_range": f"A1:{sheet.calculate_dimension()}",
            "hyperlinks": [],
            "formulas": [],
            "data_types": {},
            "merged_cells": [],
            "charts": []
        }
        
        # Analisar hyperlinks
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    sheet_info["hyperlinks"].append({
                        "cell": cell.coordinate,
                        "address": cell.hyperlink.target,
                        "text": cell.value
                    })
                    analysis["hyperlinks"].append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "address": cell.hyperlink.target,
                        "text": cell.value
                    })
        
        # Analisar fórmulas
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    sheet_info["formulas"].append({
                        "cell": cell.coordinate,
                        "formula": cell.value
                    })
                    analysis["formulas"].append({
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "formula": cell.value
                    })
        
        # Analisar tipos de dados
        data_types = {}
        for row in sheet.iter_rows(min_row=1, max_row=min(100, sheet.max_row)):
            for cell in row:
                if cell.value is not None:
                    data_type = type(cell.value).__name__
                    data_types[data_type] = data_types.get(data_type, 0) + 1
        
        sheet_info["data_types"] = data_types
        
        # Analisar células mescladas
        for merged_range in sheet.merged_cells.ranges:
            sheet_info["merged_cells"].append(str(merged_range))
        
        # Analisar gráficos
        for chart in sheet._charts:
            sheet_info["charts"].append({
                "type": type(chart).__name__,
                "title": getattr(chart, 'title', 'Sem título')
            })
            analysis["charts"].append({
                "sheet": sheet_name,
                "type": type(chart).__name__,
                "title": getattr(chart, 'title', 'Sem título')
            })
        
        analysis["sheets"].append(sheet_info)
        
        # Mostrar resumo da aba
        print(f"   📏 Dimensões: {sheet_info['dimensions']}")
        print(f"   🔗 Hyperlinks: {len(sheet_info['hyperlinks'])}")
        print(f"   🧮 Fórmulas: {len(sheet_info['formulas'])}")
        print(f"   📊 Gráficos: {len(sheet_info['charts'])}")
        print(f"   📋 Células mescladas: {len(sheet_info['merged_cells'])}")
    
    # Analisar links externos
    if hasattr(wb, 'external_links'):
        for link in wb.external_links:
            analysis["external_links"].append({
                "type": link.type,
                "target": link.target
            })
    
    # Identificar possíveis fontes de dados
    data_sources = []
    for sheet_info in analysis["sheets"]:
        for formula in sheet_info["formulas"]:
            # Procurar por referências a arquivos externos
            if re.search(r'\[.*\.xlsx?\]', formula["formula"]):
                data_sources.append({
                    "type": "Excel File Reference",
                    "source": formula["formula"],
                    "sheet": sheet_info["name"],
                    "cell": formula["cell"]
                })
            
            # Procurar por funções de dados externos
            if any(func in formula["formula"].upper() for func in ['IMPORTXML', 'IMPORTHTML', 'IMPORTFEED', 'GOOGLEFINANCE']):
                data_sources.append({
                    "type": "Web Data Function",
                    "source": formula["formula"],
                    "sheet": sheet_info["name"],
                    "cell": formula["cell"]
                })
    
    analysis["data_sources"] = data_sources
    
    # Gerar resumo
    analysis["summary"] = {
        "total_sheets": len(analysis["sheets"]),
        "total_hyperlinks": len(analysis["hyperlinks"]),
        "total_formulas": len(analysis["formulas"]),
        "total_charts": len(analysis["charts"]),
        "total_external_links": len(analysis["external_links"]),
        "total_data_sources": len(analysis["data_sources"])
    }
    
    return analysis

def print_analysis_report(analysis):
    """Imprime um relatório detalhado da análise"""
    
    print("\n" + "=" * 60)
    print("📋 RELATÓRIO COMPLETO DA ANÁLISE")
    print("=" * 60)
    
    print(f"\n📄 Arquivo: {analysis['file_name']}")
    print(f"📏 Tamanho: {analysis['file_size_mb']} MB")
    
    print(f"\n📊 RESUMO GERAL:")
    print(f"   • Total de abas: {analysis['summary']['total_sheets']}")
    print(f"   • Total de hyperlinks: {analysis['summary']['total_hyperlinks']}")
    print(f"   • Total de fórmulas: {analysis['summary']['total_formulas']}")
    print(f"   • Total de gráficos: {analysis['summary']['total_charts']}")
    print(f"   • Links externos: {analysis['summary']['total_external_links']}")
    print(f"   • Fontes de dados: {analysis['summary']['total_data_sources']}")
    
    print(f"\n📋 DETALHES POR ABA:")
    for sheet in analysis["sheets"]:
        print(f"\n   📄 {sheet['name']}")
        print(f"      📏 Dimensões: {sheet['dimensions']}")
        print(f"      🔗 Hyperlinks: {len(sheet['hyperlinks'])}")
        print(f"      🧮 Fórmulas: {len(sheet['formulas'])}")
        print(f"      📊 Gráficos: {len(sheet['charts'])}")
        
        if sheet['hyperlinks']:
            print(f"      🔗 Links encontrados:")
            for link in sheet['hyperlinks'][:5]:  # Mostrar apenas os primeiros 5
                print(f"         • {link['cell']}: {link['address']}")
            if len(sheet['hyperlinks']) > 5:
                print(f"         ... e mais {len(sheet['hyperlinks']) - 5} links")
    
    if analysis['hyperlinks']:
        print(f"\n🔗 TODOS OS HYPERLINKS:")
        for link in analysis['hyperlinks']:
            print(f"   • {link['sheet']}!{link['cell']}: {link['address']}")
    
    if analysis['data_sources']:
        print(f"\n📋 FONTES DE DADOS IDENTIFICADAS:")
        for source in analysis['data_sources']:
            print(f"   • {source['type']} em {source['sheet']}!{source['cell']}")
            print(f"     {source['source']}")
    
    if analysis['external_links']:
        print(f"\n📋 LINKS EXTERNOS:")
        for link in analysis['external_links']:
            print(f"   • {link['type']}: {link['target']}")

def save_analysis_to_json(analysis, output_file="excel_analysis.json"):
    """Salva a análise em formato JSON"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(analysis, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n💾 Análise salva em: {output_file}")

if __name__ == "__main__":
    # Analisar a planilha Decisor.xlsx
    excel_file = "Decisor.xlsx"
    
    if Path(excel_file).exists():
        analysis = analyze_excel_file(excel_file)
        if analysis:
            print_analysis_report(analysis)
            save_analysis_to_json(analysis)
    else:
        print(f"❌ Arquivo {excel_file} não encontrado!")
        print("Certifique-se de que o arquivo está no diretório atual.") 