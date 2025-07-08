import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

def quick_excel_analysis(file_path):
    """Análise rápida da planilha Excel"""
    
    print(f"🔍 Análise Rápida: {file_path}")
    print("=" * 50)
    
    try:
        # Carregar workbook
        wb = load_workbook(file_path, data_only=False)
        
        print(f"📊 Total de abas: {len(wb.sheetnames)}")
        print(f"📋 Abas encontradas: {', '.join(wb.sheetnames)}")
        
        # Analisar cada aba
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n📄 Aba: {sheet_name}")
            print(f"   📏 Dimensões: {sheet.max_row} linhas x {sheet.max_column} colunas")
            
            # Contar hyperlinks
            hyperlink_count = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.hyperlink:
                        hyperlink_count += 1
            
            print(f"   🔗 Hyperlinks: {hyperlink_count}")
            
            # Mostrar primeiros dados
            print(f"   📝 Primeiros dados:")
            for row in range(1, min(4, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, sheet.max_column + 1)):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value:
                        row_data.append(str(cell_value)[:20])
                    else:
                        row_data.append("")
                print(f"      Linha {row}: {' | '.join(row_data)}")
        
        # Verificar links externos
        if hasattr(wb, 'external_links') and wb.external_links:
            print(f"\n🔗 Links externos encontrados:")
            for link in wb.external_links:
                print(f"   • {link.target}")
        
        print(f"\n✅ Análise concluída!")
        
    except Exception as e:
        print(f"❌ Erro na análise: {e}")

if __name__ == "__main__":
    quick_excel_analysis("Decisor.xlsx") 